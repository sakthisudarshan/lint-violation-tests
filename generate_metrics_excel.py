"""
generate_metrics_excel.py
Runs pylint on sample_code/, parses JSON output, computes R27-R38 metrics,
and writes a detailed Excel report.
"""
from __future__ import annotations

import json
import math
import os
import re
import subprocess
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# 1.  RUN PYLINT
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
SAMPLE_DIR = BASE_DIR / "sample_code"
PYLINT_JSON = BASE_DIR / "pylint_output_raw.json"


def run_pylint() -> list[dict]:
    """Run pylint with JSON output and return list of violation dicts."""
    cmd = [
        sys.executable, "-m", "pylint",
        str(SAMPLE_DIR),
        "--output-format=json",
        "--load-plugins=custom_pylint_plugin",
        f"--rcfile={BASE_DIR / '.pylintrc'}",
    ]
    # Ensure the base directory is on PYTHONPATH so the custom plugin is importable
    env = os.environ.copy()
    existing_pypath = env.get("PYTHONPATH", "")
    env["PYTHONPATH"] = (str(BASE_DIR) + os.pathsep + existing_pypath).strip(os.pathsep)

    result = subprocess.run(
        cmd, capture_output=True, text=True, cwd=str(BASE_DIR), env=env
    )
    raw_json = result.stdout.strip()
    if not raw_json:
        raw_json = "[]"
    violations = json.loads(raw_json)
    PYLINT_JSON.write_text(json.dumps(violations, indent=2), encoding="utf-8")
    # Filter out the bad-plugin-value error
    violations = [v for v in violations if v.get("symbol") != "bad-plugin-value"]
    return violations


# ─────────────────────────────────────────────────────────────────────────────
# 2.  COUNT LINES OF CODE
# ─────────────────────────────────────────────────────────────────────────────
def count_loc(directory: Path) -> tuple[int, dict[str, int]]:
    """Return (total_loc, {filename: loc}) for all .py files in directory."""
    file_loc: dict[str, int] = {}
    for py_file in directory.rglob("*.py"):
        lines = py_file.read_text(encoding="utf-8-sig", errors="ignore").splitlines()
        file_loc[py_file.name] = len(lines)
    total = sum(file_loc.values())
    return total, file_loc


# ─────────────────────────────────────────────────────────────────────────────
# 3.  COUNT SUPPRESSED VIOLATIONS  (pylint: disable= inline comments)
# ─────────────────────────────────────────────────────────────────────────────
DISABLE_RE = re.compile(r"#\s*pylint\s*:\s*disable\s*=", re.IGNORECASE)

def count_suppressions(directory: Path) -> int:
    """Count inline pylint:disable comments across all .py source files."""
    count = 0
    for py_file in directory.rglob("*.py"):
        for line in py_file.read_text(encoding="utf-8-sig", errors="ignore").splitlines():
            if DISABLE_RE.search(line):
                count += 1
    return count


# ─────────────────────────────────────────────────────────────────────────────
# 4.  COUNT TOTAL DECLARED VARIABLES (simple heuristic via AST)
# ─────────────────────────────────────────────────────────────────────────────
def _safe_parse(py_file: Path):
    """Parse a .py file, handling BOM and encoding gracefully."""
    import ast
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return ast.parse(py_file.read_text(encoding=enc, errors="ignore"))
        except SyntaxError:
            continue
    return None


def count_declared_names(directory: Path) -> int:
    """Count assignment targets + function/class defs as a proxy for named identifiers."""
    import ast
    total = 0
    for py_file in directory.rglob("*.py"):
        tree = _safe_parse(py_file)
        if tree is None:
            continue
        for node in ast.walk(tree):
            if isinstance(node, (ast.Assign, ast.AnnAssign, ast.AugAssign)):
                total += 1
            elif isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
                total += 1
    return total


def count_functions(directory: Path) -> int:
    """Count all function definitions."""
    import ast
    total = 0
    for py_file in directory.rglob("*.py"):
        tree = _safe_parse(py_file)
        if tree is None:
            continue
        for node in ast.walk(tree):
            if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
                total += 1
    return total


def count_custom_rule_checks(directory: Path) -> int:
    """Count public functions (not starting with _) that the custom rule applies to."""
    import ast
    total = 0
    for py_file in directory.rglob("*.py"):
        tree = _safe_parse(py_file)
        if tree is None:
            continue
        for node in ast.walk(tree):
            if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
                if not node.name.startswith("_"):
                    total += 1
    return total


# ─────────────────────────────────────────────────────────────────────────────
# 5.  COMPUTE ALL R27–R38 METRICS
# ─────────────────────────────────────────────────────────────────────────────
def compute_metrics(violations: list[dict]) -> dict:
    """Return a dict with all raw and derived values for R27-R38."""
    total_loc, file_loc = count_loc(SAMPLE_DIR)
    total_suppressions = count_suppressions(SAMPLE_DIR)
    total_declared = count_declared_names(SAMPLE_DIR)
    total_functions = count_functions(SAMPLE_DIR)
    total_custom_checks = count_custom_rule_checks(SAMPLE_DIR)

    # ── Severity bucketing ──
    errors   = [v for v in violations if v["type"] == "error"]
    warnings = [v for v in violations if v["type"] == "warning"]
    conventions = [v for v in violations if v["type"] == "convention"]
    refactors   = [v for v in violations if v["type"] == "refactor"]

    e_count = len(errors)
    w_count = len(warnings)
    c_count = len(conventions)
    r_count = len(refactors)
    total_violations = len(violations)

    # ── Per-file violation counts ──
    file_violations: dict[str, list] = defaultdict(list)
    for v in violations:
        fname = Path(v["path"]).name
        file_violations[fname].append(v)

    files_with_violations = len(file_violations)
    files_with_3plus = sum(1 for vlist in file_violations.values() if len(vlist) >= 3)

    # ── Unique identifiers ──
    unique_rule_ids   = len({v["symbol"] for v in violations})
    unique_modules    = len({v["module"] for v in violations})
    unique_files      = len({v["path"]   for v in violations})
    unique_types      = len({v["type"]   for v in violations})

    # ── Naming / style / complexity subsets ──
    naming_viols  = [v for v in violations if v["symbol"] == "invalid-name"]
    style_viols   = [v for v in violations
                     if v["symbol"] in ("line-too-long", "trailing-whitespace",
                                        "bad-whitespace", "missing-final-newline",
                                        "trailing-newlines", "bad-indentation",
                                        "unexpected-line-ending-format",
                                        "multiple-statements")]
    complexity_viols = [v for v in violations
                        if v["symbol"] in ("too-many-branches", "too-many-statements",
                                           "too-many-nested-blocks", "too-many-arguments",
                                           "too-complex")]
    unused_viols  = [v for v in violations
                     if v["symbol"] in ("unused-variable", "unused-import",
                                        "unused-argument")]
    custom_viols  = [v for v in violations if v["symbol"] == "too-many-function-args-custom"]

    n_unused  = len(unused_viols)
    n_naming  = len(naming_viols)
    n_style   = len(style_viols)
    n_complex = len(complexity_viols)
    n_custom  = len(custom_viols)

    # ── Audit completeness (R38) ──
    expected_fields = ["type", "module", "obj", "line", "column", "path",
                       "symbol", "message", "message-id"]
    def fully_filled(v: dict) -> bool:
        return all(v.get(f) is not None and v.get(f) != "" for f in expected_fields)

    filled_violations = sum(1 for v in violations if fully_filled(v))

    # ── Line counts per violation (endLine - line + 1) for R28 proxy ──
    line_spans = []
    for v in violations:
        end = v.get("endLine") or v.get("line")
        span = (end or v["line"]) - v["line"] + 1
        line_spans.append(span)
    total_line_spans = sum(line_spans)

    # ── Violation counts per file (for R33 avg) ──
    violation_counts_per_file = [len(vl) for vl in file_violations.values()]
    avg_viols_per_file = (sum(violation_counts_per_file) / len(violation_counts_per_file)
                          if violation_counts_per_file else 0)

    # ─── Derived values (Python-specific from col10) ───────────────────────

    # R27: N = total number of JSON objects (violations) → Violation Density
    r27_raw_n    = total_violations
    r27_derived  = round((e_count * 3 + w_count * 1) / max(total_loc / 1000, 0.001), 4)
    r27_norm     = max(0, 100 - (e_count * 5 + w_count * 1))

    # R28: Total line spans = SUM(endLine - line + 1)
    r28_raw_spans = total_line_spans
    dead_alloc_pct = round((n_unused / max(total_declared, 1)) * 100, 4)
    r28_derived   = dead_alloc_pct
    r28_norm      = max(0, round(100 - (dead_alloc_pct * 50), 4))

    # R29: Semantic Consistency = 1 – (Unique Rule Types / Total Violations)
    r29_raw_naming   = n_naming
    r29_derived      = round(1 - (unique_rule_ids / max(total_violations, 1)), 4)
    convention_rate  = round((n_naming / max(total_declared, 1)) * 100, 4)
    r29_norm         = max(0, round(100 - (convention_rate * 25), 4))

    # R30: Uniformity Score = 1 – (STDEV(line spans) / MEAN(line spans))
    if len(line_spans) > 1:
        mean_spans = sum(line_spans) / len(line_spans)
        stdev_spans = math.sqrt(sum((x - mean_spans)**2 for x in line_spans) / len(line_spans))
        r30_raw_stdev  = round(stdev_spans, 4)
        r30_derived    = round(1 - (stdev_spans / max(mean_spans, 0.001)), 4)
    else:
        r30_raw_stdev = 0.0
        r30_derived   = 1.0
    style_density  = round(n_style / max(total_loc / 1000, 0.001), 4)
    r30_norm       = max(0, round(100 - (style_density * 10), 4))

    # R31: Threshold Score = COUNT(line spans > 10) / Total Violations
    long_spans     = sum(1 for s in line_spans if s > 10)
    r31_raw_complex = n_complex
    r31_derived    = round(long_spans / max(total_violations, 1), 4)
    complex_pct    = round((n_complex / max(total_functions, 1)) * 100, 4)
    r31_norm       = max(0, round(100 - (complex_pct * 10), 4))

    # R32: Impact Score = Unique Modules / Unique Files
    r32_raw_severity = {"E": e_count, "W": w_count, "C": c_count, "R": r_count}
    r32_derived      = round(unique_modules / max(unique_files, 1), 4)
    severity_ratio   = round(e_count / max(w_count + c_count + r_count, 1), 4)
    r32_norm         = max(0, 100 - (e_count * 10))

    # R33: Risk Score = Total Violations * Average Duplicate Lines
    r33_raw_hotfiles = files_with_3plus
    r33_derived      = round(total_violations * avg_viols_per_file, 4)
    hotspot_density  = round((files_with_3plus / max(files_with_violations, 1)) * 100, 4)
    r33_norm         = max(0, round(100 - (hotspot_density * 2), 4))

    # R34: Accuracy = 1 – (Unique Rule IDs – 1) / Total Violations
    r34_raw_suppress = total_suppressions
    r34_derived      = round(1 - (unique_rule_ids - 1) / max(total_violations, 1), 4)
    fp_rate          = round((total_suppressions / max(total_violations + total_suppressions, 1)) * 100, 4)
    r34_norm         = max(0, round(100 - (fp_rate * 5), 4))

    # R35: Project Enforcement = Unique Modules / Total Violations
    total_custom_passing = total_custom_checks - n_custom
    r35_raw_custom   = n_custom
    r35_derived      = round(unique_modules / max(total_violations, 1), 4)
    custom_pass_rate = round((total_custom_passing / max(total_custom_checks, 1)) * 100, 4)
    r35_norm         = custom_pass_rate

    # R36: Standardization = Unique Types / Total Violations
    config_present   = 1 if (BASE_DIR / ".pylintrc").exists() else 0
    r36_raw_config   = config_present
    r36_derived      = round(unique_types / max(total_violations, 1), 4)
    config_drift     = 0  # single dev / single config → 0 drift
    r36_norm         = max(0, 100 - (config_drift * 25))

    # R37: Gate = FAIL if errors > 0 else PASS
    gate_threshold   = 0
    r37_raw_total    = total_violations
    r37_gate         = "FAIL" if e_count > gate_threshold else "PASS"
    r37_derived      = r37_gate
    pipeline_pass_rate = 0.0 if e_count > 0 else 100.0
    r37_norm         = pipeline_pass_rate

    # R38: Audit Score = Filled Fields / Total Expected Fields
    total_expected_fields = total_violations * len(expected_fields)
    actual_filled = sum(
        sum(1 for f in expected_fields if v.get(f) is not None and v.get(f) != "")
        for v in violations
    )
    r38_raw_filled   = filled_violations
    r38_derived      = round(actual_filled / max(total_expected_fields, 1), 4)
    report_coverage  = round((filled_violations / max(total_violations, 1)) * 100, 4)
    r38_norm         = report_coverage

    return {
        "total_violations": total_violations,
        "total_loc": total_loc,
        "file_loc": file_loc,
        "e_count": e_count, "w_count": w_count,
        "c_count": c_count, "r_count": r_count,
        "unique_rule_ids": unique_rule_ids,
        "unique_modules": unique_modules,
        "unique_files": unique_files,
        "unique_types": unique_types,
        "total_functions": total_functions,
        "total_custom_checks": total_custom_checks,
        "files_with_violations": files_with_violations,
        "files_with_3plus": files_with_3plus,
        "total_suppressions": total_suppressions,
        "total_declared": total_declared,
        "avg_viols_per_file": avg_viols_per_file,
        # R27
        "r27_raw_n": r27_raw_n,
        "r27_derived": r27_derived,
        "r27_norm": r27_norm,
        # R28
        "r28_raw_spans": r28_raw_spans,
        "r28_derived": r28_derived,
        "r28_norm": r28_norm,
        # R29
        "r29_raw_naming": r29_raw_naming,
        "r29_derived": r29_derived,
        "r29_norm": r29_norm,
        # R30
        "r30_raw_stdev": r30_raw_stdev,
        "r30_derived": r30_derived,
        "r30_norm": r30_norm,
        # R31
        "r31_raw_complex": r31_raw_complex,
        "r31_derived": r31_derived,
        "r31_norm": r31_norm,
        # R32
        "r32_raw_severity": r32_raw_severity,
        "r32_derived": r32_derived,
        "r32_norm": r32_norm,
        # R33
        "r33_raw_hotfiles": r33_raw_hotfiles,
        "r33_derived": r33_derived,
        "r33_norm": r33_norm,
        # R34
        "r34_raw_suppress": r34_raw_suppress,
        "r34_derived": r34_derived,
        "r34_norm": r34_norm,
        # R35
        "r35_raw_custom": r35_raw_custom,
        "r35_derived": r35_derived,
        "r35_norm": r35_norm,
        # R36
        "r36_raw_config": r36_raw_config,
        "r36_derived": r36_derived,
        "r36_norm": r36_norm,
        # R37
        "r37_raw_total": r37_raw_total,
        "r37_derived": r37_derived,
        "r37_norm": r37_norm,
        # R38
        "r38_raw_filled": r38_raw_filled,
        "r38_derived": r38_derived,
        "r38_norm": r38_norm,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 6.  BUILD EXCEL REPORT
# ─────────────────────────────────────────────────────────────────────────────
METRIC_ROWS = [
    {
        "row_id": "R27",
        "l1": "White Box",
        "l2": "Static Code Analysis",
        "l3": "Lint / Rule Violations",
        "l4": "Rule Detection Test",
        "l5": "Violation Density per KLOC",
        "description": "Count of lint violations bucketed by severity (error, warning, info) per 1000 lines of code",
        "tool_primary": "pylint",
        "tool_secondary": "flake8",
        "emitted_directly": "No",
        "python_derivation": "N = total number of JSON objects (violations)",
        "raw_formula": "Violation Density = (Errors×3 + Warnings×1) / (LOC/1000)",
        "threshold": "0 blocking errors; < 10 warnings per KLOC",
        "norm_formula": "MAX(0, 100 – (Errors×5 + Warnings×1))",
        "frequency": "Every Commit / PR",
        "raw_key": "r27_raw_n",
        "raw_label": "Total Violations (N)",
        "derived_key": "r27_derived",
        "derived_label": "Violation Density Score",
        "norm_key": "r27_norm",
    },
    {
        "row_id": "R28",
        "l1": "White Box",
        "l2": "Static Code Analysis",
        "l3": "Lint / Rule Violations",
        "l4": "Unused Variable Detection",
        "l5": "Resource Waste Identification",
        "description": "Scans for variables declared but never referenced; measures technical debt from dead allocations.",
        "tool_primary": "pylint",
        "tool_secondary": "flake8",
        "emitted_directly": "No",
        "python_derivation": "Total Line Spans = SUM(endLine – line + 1)",
        "raw_formula": "Dead Allocation % = (Unused Variables / Total Declared Variables) × 100",
        "threshold": "< 1% unused variable declarations per module",
        "norm_formula": "MAX(0, 100 – (Dead_Allocation% × 50))",
        "frequency": "Every Commit / PR",
        "raw_key": "r28_raw_spans",
        "raw_label": "Total Violation Line Spans",
        "derived_key": "r28_derived",
        "derived_label": "Dead Allocation %",
        "norm_key": "r28_norm",
    },
    {
        "row_id": "R29",
        "l1": "White Box",
        "l2": "Static Code Analysis",
        "l3": "Lint / Rule Violations",
        "l4": "Naming Convention Validation",
        "l5": "Semantic Consistency Score",
        "description": "Verifies variables, functions, classes follow casing standards; measures guessability of code.",
        "tool_primary": "pylint",
        "tool_secondary": "flake8",
        "emitted_directly": "No",
        "python_derivation": "Semantic Consistency = 1 – (Unique Rule Types / Total Violations)",
        "raw_formula": "Convention Violation Rate = (Naming Violations / Total Named Identifiers) × 100",
        "threshold": "< 2% naming convention violations across codebase",
        "norm_formula": "MAX(0, 100 – (Convention_Violation_Rate × 25))",
        "frequency": "Every Commit / PR",
        "raw_key": "r29_raw_naming",
        "raw_label": "Naming Violations (C0103)",
        "derived_key": "r29_derived",
        "derived_label": "Semantic Consistency Score",
        "norm_key": "r29_norm",
    },
    {
        "row_id": "R30",
        "l1": "White Box",
        "l2": "Static Code Analysis",
        "l3": "Lint / Rule Violations",
        "l4": "Code Style Rule Validation",
        "l5": "Syntactic Uniformity Score",
        "description": "Enforces indentation, line length, whitespace rules; ensures visual consistency across developers.",
        "tool_primary": "pylint",
        "tool_secondary": "flake8",
        "emitted_directly": "No",
        "python_derivation": "Uniformity Score = 1 – (STDEV(violation line spans) / MEAN(violation line spans))",
        "raw_formula": "Style Violation Density = Style Violations / (LOC / 1000)",
        "threshold": "< 5 style violations per KLOC",
        "norm_formula": "MAX(0, 100 – (Style_Violation_Density × 10))",
        "frequency": "Every Commit / PR",
        "raw_key": "r30_raw_stdev",
        "raw_label": "STDEV of Violation Line Spans",
        "derived_key": "r30_derived",
        "derived_label": "Syntactic Uniformity Score",
        "norm_key": "r30_norm",
    },
    {
        "row_id": "R31",
        "l1": "White Box",
        "l2": "Static Code Analysis",
        "l3": "Lint / Rule Violations",
        "l4": "Complexity Rule Detection",
        "l5": "Structural Threshold Monitoring",
        "description": "Flags functions exceeding nesting/length limits; measures mental load required to process a module.",
        "tool_primary": "pylint",
        "tool_secondary": "flake8",
        "emitted_directly": "No",
        "python_derivation": "Threshold Score = COUNT(line spans > 10) / Total Violations",
        "raw_formula": "Complexity Hotspot % = (Functions Exceeding Threshold / Total Functions) × 100",
        "threshold": "< 10% of functions exceed complexity thresholds",
        "norm_formula": "MAX(0, 100 – (Complexity_Hotspot% × 10))",
        "frequency": "Every Commit / PR",
        "raw_key": "r31_raw_complex",
        "raw_label": "Complexity Violations (R0912/R0913/R0915/R1702)",
        "derived_key": "r31_derived",
        "derived_label": "Threshold Score",
        "norm_key": "r31_norm",
    },
    {
        "row_id": "R32",
        "l1": "White Box",
        "l2": "Static Code Analysis",
        "l3": "Lint / Rule Violations",
        "l4": "Rule Severity Classification",
        "l5": "Impact Prioritization",
        "description": "Categorizes violations as Errors/Warnings/Info; helps focus on critical structural flaws first.",
        "tool_primary": "pylint",
        "tool_secondary": "flake8",
        "emitted_directly": "No",
        "python_derivation": "Impact Score = Unique Modules / Unique Files",
        "raw_formula": "Severity Ratio = Error Violations / (Warnings + Conventions + Refactors)",
        "threshold": "0 Error-level violations in production build",
        "norm_formula": "MAX(0, 100 – (Error_Count × 10))",
        "frequency": "Every Commit / PR",
        "raw_key": "r32_raw_severity",
        "raw_label": "Severity Breakdown (E/W/C/R)",
        "derived_key": "r32_derived",
        "derived_label": "Impact Score (Unique Modules / Unique Files)",
        "norm_key": "r32_norm",
    },
    {
        "row_id": "R33",
        "l1": "White Box",
        "l2": "Static Code Analysis",
        "l3": "Lint / Rule Violations",
        "l4": "Multiple Violations Detection",
        "l5": "Aggregated Risk Assessment",
        "description": "Identifies modules with high density of different rule breaks; measures system instability.",
        "tool_primary": "pylint",
        "tool_secondary": "flake8",
        "emitted_directly": "No",
        "python_derivation": "Risk Score = Total Violations × Average Violations per File",
        "raw_formula": "Hotspot Density = (Files with ≥ 3 violations / Total Files with violations) × 100",
        "threshold": "< 30% of files classified as hotspots",
        "norm_formula": "MAX(0, 100 – (Hotspot_Density × 2))",
        "frequency": "Every Commit / PR",
        "raw_key": "r33_raw_hotfiles",
        "raw_label": "Files with ≥3 Violations",
        "derived_key": "r33_derived",
        "derived_label": "Aggregated Risk Score",
        "norm_key": "r33_norm",
    },
    {
        "row_id": "R34",
        "l1": "White Box",
        "l2": "Static Code Analysis",
        "l3": "Lint / Rule Violations",
        "l4": "False Positive Prevention",
        "l5": "Accuracy Tuning",
        "description": "Filters intentional suppressions; measures reliability of analysis to prevent warning fatigue.",
        "tool_primary": "pylint",
        "tool_secondary": "flake8",
        "emitted_directly": "No",
        "python_derivation": "Accuracy = 1 – (Unique Rule IDs – 1) / Total Violations",
        "raw_formula": "False Positive Rate % = (Suppressed Violations / Total Flagged) × 100",
        "threshold": "< 10% suppression rate (high suppression → misconfigured rules)",
        "norm_formula": "MAX(0, 100 – (False_Positive_Rate% × 5))",
        "frequency": "Every Commit / PR",
        "raw_key": "r34_raw_suppress",
        "raw_label": "Inline pylint:disable Count",
        "derived_key": "r34_derived",
        "derived_label": "Accuracy Score",
        "norm_key": "r34_norm",
    },
    {
        "row_id": "R35",
        "l1": "White Box",
        "l2": "Static Code Analysis",
        "l3": "Lint / Rule Violations",
        "l4": "Custom Rule Validation",
        "l5": "Project-Specific Enforcement",
        "description": "Enforces unique project rules; measures compliance with internal architecture standards.",
        "tool_primary": "pylint",
        "tool_secondary": "flake8",
        "emitted_directly": "No",
        "python_derivation": "Project Enforcement = Unique Modules / Total Violations",
        "raw_formula": "Custom Rule Pass Rate % = (Custom Rule Checks Passing / Total Custom Rule Checks) × 100",
        "threshold": "100% of project-specific custom rules passing",
        "norm_formula": "Score = Custom Rule Pass Rate % [gate at 100%]",
        "frequency": "Every Commit / PR",
        "raw_key": "r35_raw_custom",
        "raw_label": "Custom Rule Violations (W9001)",
        "derived_key": "r35_derived",
        "derived_label": "Project Enforcement Score",
        "norm_key": "r35_norm",
    },
    {
        "row_id": "R36",
        "l1": "White Box",
        "l2": "Static Code Analysis",
        "l3": "Lint / Rule Violations",
        "l4": "Configuration File Handling",
        "l5": "Environment Standardization",
        "description": "Manages config files for active rules; ensures team-wide quality setting consistency.",
        "tool_primary": "pylint",
        "tool_secondary": "flake8",
        "emitted_directly": "No",
        "python_derivation": "Standardization = Unique Violation Types / Total Violations",
        "raw_formula": "Config Drift Score = Count(Devs with lint config deviating from team standard)",
        "threshold": "0 developers with non-standard lint configuration",
        "norm_formula": "MAX(0, 100 – (Config_Drift_Score × 25))",
        "frequency": "Every Commit / PR",
        "raw_key": "r36_raw_config",
        "raw_label": ".pylintrc Config File Present (1=Yes, 0=No)",
        "derived_key": "r36_derived",
        "derived_label": "Standardization Score (Unique Types / Total Violations)",
        "norm_key": "r36_norm",
    },
    {
        "row_id": "R37",
        "l1": "White Box",
        "l2": "Static Code Analysis",
        "l3": "Lint / Rule Violations",
        "l4": "CI/CD Integration Validation",
        "l5": "Automated Gatekeeping",
        "description": "Runs quality checks during build; blocks code failing minimum quality bar from merge.",
        "tool_primary": "pylint",
        "tool_secondary": "flake8",
        "emitted_directly": "No",
        "python_derivation": 'Gate = IF(Total Violations > Threshold, "FAIL", "PASS")',
        "raw_formula": "Pipeline Gate Pass Rate % = (Builds passing lint gate / Total Builds) × 100",
        "threshold": "100% of builds pass lint quality gate before merge",
        "norm_formula": "Score = Pipeline Gate Pass Rate % [gate at 100%]",
        "frequency": "Every Commit / PR",
        "raw_key": "r37_raw_total",
        "raw_label": "Total Violations (gate input)",
        "derived_key": "r37_derived",
        "derived_label": 'Gate Result (PASS if errors=0)',
        "norm_key": "r37_norm",
    },
    {
        "row_id": "R38",
        "l1": "White Box",
        "l2": "Static Code Analysis",
        "l3": "Lint / Rule Violations",
        "l4": "Violation Reporting Validation",
        "l5": "Quality Audit Trail",
        "description": "Generates detailed logs of all issues with line numbers; provides transparent record of code health.",
        "tool_primary": "pylint",
        "tool_secondary": "flake8",
        "emitted_directly": "No",
        "python_derivation": "Audit Score = Filled Fields / Total Expected Fields",
        "raw_formula": "Report Coverage % = (Violations with complete audit fields / Total Violations) × 100",
        "threshold": ">= 95% of violations have complete audit log entry",
        "norm_formula": "Score = Report Coverage % [gate at 95%]",
        "frequency": "Every Commit / PR",
        "raw_key": "r38_raw_filled",
        "raw_label": "Violations with All Fields Filled",
        "derived_key": "r38_derived",
        "derived_label": "Audit Completeness Ratio",
        "norm_key": "r38_norm",
    },
]


def score_to_band(score) -> str:
    """Convert a 0-100 numeric score to a quality band label."""
    if isinstance(score, str):  # e.g. "PASS"/"FAIL"
        return "PASS ✓" if score == "PASS" else "FAIL ✗"
    if score >= 90:
        return "Excellent"
    if score >= 75:
        return "Good"
    if score >= 50:
        return "Moderate"
    if score >= 25:
        return "Poor"
    return "Critical"


def score_fill(score) -> PatternFill:
    if isinstance(score, str):
        color = "92D050" if score == "PASS" else "FF0000"
        return PatternFill("solid", fgColor=color)
    if score >= 90:
        return PatternFill("solid", fgColor="92D050")   # green
    if score >= 75:
        return PatternFill("solid", fgColor="FFFF00")   # yellow
    if score >= 50:
        return PatternFill("solid", fgColor="FFC000")   # orange
    if score >= 25:
        return PatternFill("solid", fgColor="FF6600")   # dark orange
    return PatternFill("solid", fgColor="FF0000")        # red


def build_excel(metrics: dict, violations: list[dict], output_path: Path):
    wb = openpyxl.Workbook()

    # ── colour palette ──
    DARK_BLUE  = "1F3864"
    MID_BLUE   = "2E75B6"
    LIGHT_BLUE = "D9E1F2"
    HEADER_BG  = "17375E"
    ALT_ROW    = "EBF3FB"
    WHITE      = "FFFFFF"
    TITLE_FG   = "FFFFFF"
    BLACK      = "000000"

    def hdr_font(bold=True, color=TITLE_FG, size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def body_font(bold=False, color=BLACK, size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def thin_border():
        side = Side(style="thin", color="BFBFBF")
        return Border(left=side, right=side, top=side, bottom=side)

    def med_border():
        side = Side(style="medium", color="2E75B6")
        return Border(left=side, right=side, top=side, bottom=side)

    wrap_align  = Alignment(wrap_text=True, vertical="top", horizontal="left")
    center_mid  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1 – SUMMARY DASHBOARD
    # ═══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Metrics Dashboard"
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:Q2")
    title_cell = ws["A1"]
    title_cell.value = (
        "TESTABLE  |  Lint / Rule Violation Metrics Report  |  "
        "Python (pylint)  |  R27-R38  |  White Box - Static Code Analysis"
    )
    title_cell.font = Font(bold=True, color=TITLE_FG, size=14, name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    title_cell.alignment = center_mid

    # Sub-title / meta
    ws.merge_cells("A3:Q3")
    sub_cell = ws["A3"]
    sub_cell.value = (
        f"Tool: pylint  |  Scope: sample_code/  |  "
        f"Total LOC: {metrics['total_loc']}  |  "
        f"Total Violations: {metrics['total_violations']}  |  "
        f"E={metrics['e_count']}  W={metrics['w_count']}  "
        f"C={metrics['c_count']}  R={metrics['r_count']}"
    )
    sub_cell.font   = Font(bold=False, color=TITLE_FG, size=10, name="Calibri")
    sub_cell.fill   = PatternFill("solid", fgColor=MID_BLUE)
    sub_cell.alignment = center_mid

    # Column headers  (row 5)
    HEADERS = [
        "Row ID", "L1 Strategy", "L2 Testing Type", "L3 Technique",
        "L4 Classification", "L5 Metric",
        "What Does The Metric Measure?",
        "Tool (Primary)", "Tool (Secondary)", "Metric Emitted Directly?",
        "Python Derivation Formula",
        "Raw Measurement Formula",
        "Expected Value / Threshold",
        "Normalisation Formula (0–100)",
        "Execution Frequency",
        "Raw Value\n(from pylint output)",
        "Raw Value Label",
        "Derived Value\n(Python formula applied)",
        "Derived Value Label",
        "Normalised Score\n(0–100)",
        "Quality Band",
    ]

    COL_WIDTHS = [
        8, 14, 22, 24, 28, 30,
        42,
        14, 14, 14,
        52,
        52,
        40,
        40,
        18,
        18, 28,
        18, 36,
        16, 14,
    ]

    for col_idx, (hdr, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=5, column=col_idx, value=hdr)
        cell.font = hdr_font(size=9)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = center_mid
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[5].height = 42

    # Data rows
    for r_idx, mrow in enumerate(METRIC_ROWS):
        row_num = 6 + r_idx
        fill = PatternFill("solid", fgColor=ALT_ROW if r_idx % 2 == 0 else WHITE)

        raw_val = metrics[mrow["raw_key"]]
        derived_val = metrics[mrow["derived_key"]]
        norm_val    = metrics[mrow["norm_key"]]

        # Format severity dict nicely
        if isinstance(raw_val, dict):
            raw_display = "  |  ".join(f"{k}={v}" for k, v in raw_val.items())
        else:
            raw_display = raw_val

        band = score_to_band(norm_val)
        band_fill = score_fill(norm_val)

        row_data = [
            mrow["row_id"],
            mrow["l1"],
            mrow["l2"],
            mrow["l3"],
            mrow["l4"],
            mrow["l5"],
            mrow["description"],
            mrow["tool_primary"],
            mrow["tool_secondary"],
            mrow["emitted_directly"],
            mrow["python_derivation"],
            mrow["raw_formula"],
            mrow["threshold"],
            mrow["norm_formula"],
            mrow["frequency"],
            raw_display,
            mrow["raw_label"],
            derived_val,
            mrow["derived_label"],
            norm_val,
            band,
        ]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font   = body_font(bold=(col_idx == 1))
            cell.fill   = band_fill if col_idx == 21 else fill
            cell.alignment = wrap_align if col_idx not in (1, 8, 9, 10, 15, 16, 20, 21) else center_mid
            cell.border = thin_border()

        ws.row_dimensions[row_num].height = 75

    ws.freeze_panes = "A6"

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2 – RAW PYLINT VIOLATIONS
    # ═══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Raw Pylint Violations")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:K1")
    t2 = ws2["A1"]
    t2.value = f"Pylint Raw Output  |  Total Violations: {len(violations)}  |  Source: sample_code/"
    t2.font  = Font(bold=True, color=TITLE_FG, size=12, name="Calibri")
    t2.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    t2.alignment = center_mid

    raw_headers = ["#", "Type", "Severity Code", "Module", "Object",
                   "File", "Line", "Column", "Symbol (Rule ID)", "Message", "Message-ID"]
    raw_widths  = [6, 12, 14, 30, 25, 35, 8, 8, 30, 60, 14]
    for ci, (h, w) in enumerate(zip(raw_headers, raw_widths), 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font = hdr_font(size=9)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = center_mid
        cell.border = thin_border()
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.row_dimensions[2].height = 28

    TYPE_COLORS = {
        "error":      "FFB3B3",
        "warning":    "FFE0A3",
        "convention": "D9EAD3",
        "refactor":   "CFE2F3",
    }
    for vi, v in enumerate(violations, 1):
        rn = vi + 2
        row_fill = PatternFill("solid", fgColor=TYPE_COLORS.get(v["type"], "F2F2F2"))
        row_vals = [
            vi,
            v["type"].upper()[0],
            v["type"].capitalize(),
            v.get("module", ""),
            v.get("obj", ""),
            v.get("path", ""),
            v.get("line", ""),
            v.get("column", ""),
            v.get("symbol", ""),
            v.get("message", ""),
            v.get("message-id", ""),
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=rn, column=ci, value=val)
            cell.font   = body_font(size=9)
            cell.fill   = row_fill
            cell.border = thin_border()
            cell.alignment = wrap_align if ci in (9, 10) else center_mid
        ws2.row_dimensions[rn].height = 28

    ws2.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 3 – METRIC COMPUTATION DETAIL
    # ═══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Computation Detail")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:E1")
    t3 = ws3["A1"]
    t3.value = "Metric Computation Detail  |  Inputs & Step-by-Step Derivation"
    t3.font  = Font(bold=True, color=TITLE_FG, size=12, name="Calibri")
    t3.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    t3.alignment = center_mid

    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 55
    ws3.column_dimensions["D"].width = 25
    ws3.column_dimensions["E"].width = 20

    det_headers = ["Row ID", "Parameter", "Formula / Description", "Value", "Unit"]
    for ci, h in enumerate(det_headers, 1):
        cell = ws3.cell(row=2, column=ci, value=h)
        cell.font = hdr_font(size=9)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = center_mid
        cell.border = thin_border()
    ws3.row_dimensions[2].height = 28

    det_rows = [
        # Global inputs
        ("GLOBAL", "Total LOC",         "Lines of code across all sample_code/*.py files",       metrics["total_loc"],                "lines"),
        ("GLOBAL", "Total Violations",  "Number of JSON objects in pylint --output-format=json",  metrics["total_violations"],         "count"),
        ("GLOBAL", "Error (E) Count",   "violations where type == 'error'",                       metrics["e_count"],                   "count"),
        ("GLOBAL", "Warning (W) Count", "violations where type == 'warning'",                     metrics["w_count"],                   "count"),
        ("GLOBAL", "Convention (C)",     "violations where type == 'convention'",                  metrics["c_count"],                  "count"),
        ("GLOBAL", "Refactor (R) Count","violations where type == 'refactor'",                    metrics["r_count"],                   "count"),
        ("GLOBAL", "Unique Rule IDs",   "len({v['symbol'] for v in violations})",                 metrics["unique_rule_ids"],           "count"),
        ("GLOBAL", "Unique Modules",    "len({v['module'] for v in violations})",                 metrics["unique_modules"],            "count"),
        ("GLOBAL", "Unique Files",      "len({v['path'] for v in violations})",                   metrics["unique_files"],              "count"),
        ("GLOBAL", "Unique Types",      "len({v['type'] for v in violations})",                   metrics["unique_types"],              "count"),
        ("GLOBAL", "Total Functions",   "AST count of FunctionDef nodes in sample_code/",         metrics["total_functions"],           "count"),
        ("GLOBAL", "Total Declared",    "AST count of assignment + def + class nodes",            metrics["total_declared"],            "count"),
        ("GLOBAL", "Suppressions",      "pylint:disable inline comment count",                    metrics["total_suppressions"],        "count"),
        ("GLOBAL", "Files w Violations","len({v['path'] for v in violations})",                   metrics["files_with_violations"],     "files"),
        ("GLOBAL", "Hotfiles (≥3 viols)","Files with 3 or more violations",                      metrics["files_with_3plus"],          "files"),
        ("GLOBAL", "Avg viols/file",    "total_violations / files_with_violations",               round(metrics["avg_viols_per_file"],3),"float"),
        ("GLOBAL", "Custom Checks",     "Public functions (not starting with _)",                 metrics["total_custom_checks"],       "count"),
        # R27
        ("R27",    "Raw N",             "total_violations (all JSON objects)",                    metrics["r27_raw_n"],                 "count"),
        ("R27",    "Violation Density", "(E×3 + W×1) / (LOC/1000)",                             metrics["r27_derived"],               "density"),
        ("R27",    "Norm Score",        "MAX(0, 100 – (E×5 + W×1))",                             metrics["r27_norm"],                  "0-100"),
        # R28
        ("R28",    "Raw Line Spans",    "SUM(endLine – line + 1) per violation",                 metrics["r28_raw_spans"],             "lines"),
        ("R28",    "Dead Alloc %",      "(unused_viols / total_declared) × 100",                 metrics["r28_derived"],               "%"),
        ("R28",    "Norm Score",        "MAX(0, 100 – (Dead_Alloc% × 50))",                      metrics["r28_norm"],                  "0-100"),
        # R29
        ("R29",    "Naming Violations", "count where symbol == 'invalid-name'",                  metrics["r29_raw_naming"],            "count"),
        ("R29",    "Semantic Consist.", "1 – (unique_rule_ids / total_violations)",               metrics["r29_derived"],              "ratio"),
        ("R29",    "Norm Score",        "MAX(0, 100 – (conv_rate × 25))",                        metrics["r29_norm"],                  "0-100"),
        # R30
        ("R30",    "STDEV Line Spans",  "STDEV of (endLine – line + 1) per violation",           metrics["r30_raw_stdev"],             "lines"),
        ("R30",    "Uniformity Score",  "1 – (STDEV / MEAN) of violation line spans",            metrics["r30_derived"],               "ratio"),
        ("R30",    "Norm Score",        "MAX(0, 100 – (style_density × 10))",                    metrics["r30_norm"],                  "0-100"),
        # R31
        ("R31",    "Complex Violations","R0912/R0913/R0915/R1702 count",                         metrics["r31_raw_complex"],           "count"),
        ("R31",    "Threshold Score",   "COUNT(line_span > 10) / total_violations",              metrics["r31_derived"],               "ratio"),
        ("R31",    "Norm Score",        "MAX(0, 100 – (complex_pct × 10))",                      metrics["r31_norm"],                  "0-100"),
        # R32
        ("R32",    "E/W/C/R Counts",    "Violations by severity type",
         f"E={metrics['e_count']} W={metrics['w_count']} C={metrics['c_count']} R={metrics['r_count']}", "mixed"),
        ("R32",    "Impact Score",      "unique_modules / unique_files",                         metrics["r32_derived"],               "ratio"),
        ("R32",    "Norm Score",        "MAX(0, 100 – (E_count × 10))",                          metrics["r32_norm"],                  "0-100"),
        # R33
        ("R33",    "Hotfiles",          "Files with ≥ 3 violations",                             metrics["r33_raw_hotfiles"],          "count"),
        ("R33",    "Risk Score",        "total_violations × avg_viols_per_file",                 metrics["r33_derived"],               "score"),
        ("R33",    "Norm Score",        "MAX(0, 100 – (hotspot_density × 2))",                   metrics["r33_norm"],                  "0-100"),
        # R34
        ("R34",    "Suppressions",      "inline pylint:disable= comment count",                  metrics["r34_raw_suppress"],          "count"),
        ("R34",    "Accuracy Score",    "1 – (unique_rule_ids – 1) / total_violations",          metrics["r34_derived"],               "ratio"),
        ("R34",    "Norm Score",        "MAX(0, 100 – (fp_rate × 5))",                           metrics["r34_norm"],                  "0-100"),
        # R35
        ("R35",    "Custom Violations", "count where symbol == 'too-many-function-args-custom'", metrics["r35_raw_custom"],            "count"),
        ("R35",    "Enforcement Score", "unique_modules / total_violations",                     metrics["r35_derived"],               "ratio"),
        ("R35",    "Norm Score",        "Custom Rule Pass Rate %",                               metrics["r35_norm"],                  "0-100"),
        # R36
        ("R36",    ".pylintrc Present", "1 if .pylintrc exists else 0",                          metrics["r36_raw_config"],            "bool"),
        ("R36",    "Standardization",   "unique_types / total_violations",                       metrics["r36_derived"],               "ratio"),
        ("R36",    "Norm Score",        "MAX(0, 100 – (config_drift × 25))",                     metrics["r36_norm"],                  "0-100"),
        # R37
        ("R37",    "Total Violations",  "pylint total violations (gate input)",                  metrics["r37_raw_total"],             "count"),
        ("R37",    "Gate Result",       "FAIL if errors > 0 else PASS",                          metrics["r37_derived"],               "PASS/FAIL"),
        ("R37",    "Norm Score",        "100 if PASS else 0",                                    metrics["r37_norm"],                  "0-100"),
        # R38
        ("R38",    "Complete Entries",  "violations where all 9 fields are non-null",            metrics["r38_raw_filled"],            "count"),
        ("R38",    "Audit Ratio",       "filled_fields / (total_violations × 9)",               metrics["r38_derived"],               "ratio"),
        ("R38",    "Norm Score",        "report_coverage % (gate ≥ 95%)",                        metrics["r38_norm"],                  "0-100"),
    ]

    ROW_COLORS = {
        "GLOBAL": "E2EFDA",
        "R27": "DDEEFF", "R28": "E8F4FD", "R29": "FFF2CC",
        "R30": "FCE4D6", "R31": "EAD1DC", "R32": "D9EAD3",
        "R33": "CFE2F3", "R34": "FFF3CD", "R35": "D4E6F1",
        "R36": "FDEBD0", "R37": "FADBD8", "R38": "D5F5E3",
    }

    for ri, (rid, param, formula, value, unit) in enumerate(det_rows, start=3):
        row_fill = PatternFill("solid", fgColor=ROW_COLORS.get(rid, "FFFFFF"))
        for ci, val in enumerate([rid, param, formula, value, unit], 1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.font   = body_font(bold=(ci == 1), size=9)
            cell.fill   = row_fill
            cell.border = thin_border()
            cell.alignment = wrap_align if ci == 3 else center_mid
        ws3.row_dimensions[ri].height = 22

    ws3.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 4 – FILE-LEVEL SUMMARY
    # ═══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("File-Level Summary")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:G1")
    t4 = ws4["A1"]
    t4.value = "File-Level Violation Summary  |  per sample_code/*.py"
    t4.font  = Font(bold=True, color=TITLE_FG, size=12, name="Calibri")
    t4.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    t4.alignment = center_mid

    file_headers = ["File", "LOC", "Total Violations", "E", "W", "C", "R", "Hotspot?"]
    file_widths  = [40, 8, 18, 8, 8, 8, 8, 12]
    for ci, (h, w) in enumerate(zip(file_headers, file_widths), 1):
        cell = ws4.cell(row=2, column=ci, value=h)
        cell.font = hdr_font(size=9)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = center_mid
        cell.border = thin_border()
        ws4.column_dimensions[get_column_letter(ci)].width = w

    # Build per-file data
    file_data: dict[str, dict] = {}
    for v in violations:
        fname = v["path"]
        if fname not in file_data:
            file_data[fname] = {"E": 0, "W": 0, "C": 0, "R": 0, "total": 0}
        file_data[fname]["total"] += 1
        type_map = {"error": "E", "warning": "W", "convention": "C", "refactor": "R"}
        file_data[fname][type_map.get(v["type"], "?")] += 1

    _, file_loc = count_loc(SAMPLE_DIR)

    for ri, (fpath, fd) in enumerate(sorted(file_data.items()), start=3):
        fname = Path(fpath).name
        loc   = file_loc.get(fname, 0)
        is_hot = fd["total"] >= 3
        row_fill = PatternFill("solid", fgColor="FFB3B3" if is_hot else "F2F2F2")
        row_vals = [fpath, loc, fd["total"], fd["E"], fd["W"], fd["C"], fd["R"],
                    "HOT 🔥" if is_hot else "OK"]
        for ci, val in enumerate(row_vals, 1):
            cell = ws4.cell(row=ri, column=ci, value=val)
            cell.font   = body_font(size=9)
            cell.fill   = row_fill
            cell.border = thin_border()
            cell.alignment = wrap_align if ci == 1 else center_mid
        ws4.row_dimensions[ri].height = 22

    ws4.freeze_panes = "A3"

    wb.save(str(output_path))
    print(f"\n✅  Excel report saved → {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    print(">> Running pylint ...")
    violations = run_pylint()
    print(f"   {len(violations)} violations found.")

    print(">> Computing metrics R27-R38 ...")
    metrics = compute_metrics(violations)

    print(">> Building Excel report ...")
    out_file = BASE_DIR / "Pylint_Lint_Violation_Metrics_R27_R38.xlsx"
    build_excel(metrics, violations, out_file)

    # Print summary table to console
    print("\n" + "="*90)
    print(f"{'Row':5} {'L5 Metric':35} {'Raw Value':>12} {'Derived':>12} {'Score':>8} {'Band':>12}")
    print("-"*90)
    for mrow in METRIC_ROWS:
        rv  = metrics[mrow["raw_key"]]
        dv  = metrics[mrow["derived_key"]]
        nv  = metrics[mrow["norm_key"]]
        rv_s = str(rv)[:10] if not isinstance(rv, dict) else "E/W/C/R"
        dv_s = str(dv)[:10]
        nv_s = str(nv)[:8]
        band = score_to_band(nv)
        print(f"{mrow['row_id']:5} {mrow['l5'][:35]:35} {rv_s:>12} {dv_s:>12} {nv_s:>8} {band:>12}")
    print("="*90)
